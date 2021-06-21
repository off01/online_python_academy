from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Book1.xlsx')
ws = wb['Test Scripts']

#print(wb['Test Scripts'])

for row in range(2, 11):
    for col in range(1, 6):
        char = get_column_letter(col)
        if ws[char + str(row)].value != None:
        #print(char)
            print(ws[char + str(row)].value)